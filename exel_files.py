from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from os.path import exists
from os import remove, mkdir
from constants import start_dir, sl
from log_files import info


def set_border(ws, cell_range, row, row_new):
    thin = Side(border_style='thin', color="000000")
    bold = Side(border_style='medium', color="000000")
    n = row_new
    k = ws.max_column
    s = row
    for rows in ws[cell_range]:
        a = 1
        for cell in rows:
            if s == 1:
                pass
            elif s == 2:
                cell.border = Border(top=bold, left=bold, right=bold, bottom=bold)
                cell.font = Font(bold='bold')
            elif a == 1 and s == n:
                cell.border = Border(top=thin, left=bold, right=thin, bottom=bold)
            elif a == 1:
                cell.border = Border(top=thin, left=bold, right=thin, bottom=thin)
            elif a == k and s == n:
                cell.border = Border(top=thin, left=thin, right=bold, bottom=bold)
            elif a == k:
                cell.border = Border(top=thin, left=thin, right=bold, bottom=thin)
            elif s == n:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=bold)
            else:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(horizontal='center')
            a += 1
        s += 1


def add_new_sheet(book, sheet_name, avd_time, time_start, time_finish, add_data, new_sheet=None):
    n = len(book.sheetnames)

    if not new_sheet and n == 1:
        sheet_name = f'{sheet_name}, {n}'
        sheet = book.active
        sheet.title = sheet_name
    else:
        sheet_name = f'{sheet_name}, {n + 1}'
        sheet = book.create_sheet(sheet_name)
    sheet.merge_cells('A1:F1')
    sheet['A1'] = f'Осредненение данных за {avd_time} мин. с "{time_start}" по "{time_finish}".'
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet['A1'].font = Font(bold='bold', italic=True, size=12)
    sheet.append(['Time_obs', 'Sourse_id', 'Measurand_id', # 'Measurand_label',
                  'Method_processing', 'Value', 'Count'])
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 18
    # sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['D'].width = 23
    sheet.column_dimensions['E'].width = 11
    sheet.column_dimensions['F'].width = 11
    sheet.row_dimensions[1].font = Font(bold=True)
    # else:
    #     sheet = book[f'{sheet_name}, {n}']
    filters = sheet.auto_filter
    info(f'Добавлен новый лист "{sheet_name}" в книге.', start_dir)
    return book, sheet, filters


def add_exel_files(data: list, avd_time: int, time_start, time_finish, time_begin=None, time_end=None):

    sheet_name = f'AVG time {avd_time} minute'
    dir_name = f'{start_dir}{sl}ExelFiles'
    file_name = f'{dir_name}{sl}{time_start}_{time_finish} {sheet_name}.xlsx'
    # book = None
    add_data = 0
    if not exists(dir_name):
        mkdir(dir_name)
    else:
        if exists(file_name):
            remove(file_name)
    book = Workbook()
    # if not book:
    #     book = Workbook()

    # if not exists(dir_name):
    #     mkdir(dir_name)

    # else:
    #     if exists(file_name):
    #         z = 0
    #         while z <= 5:
    #             try:
    #                 book = load_workbook(file_name)
    #                 add_data = 1
    #                 break
    #             except OSError:
    #                 info(f'Ошибка открытия файла "{file_name}"', start_dir)
    #                 z += 1
    #                 if z > 5:
    #                     return
    #                 sleep(1/2)
    #     else:
    #         book = Workbook()

    book, sheet, filters = add_new_sheet(
        book=book, sheet_name=sheet_name, avd_time=avd_time, time_start=time_start,
        time_finish=time_finish, add_data=add_data)

    row = sheet.max_row
    row_new = row
    # if row >= 1048576:  # 1048576
    #     book, sheet, filters = add_new_sheet(
    #         book=book, sheet_name=sheet_name, avd_time=avd_time, time_start=time_start,
    #         time_finish=time_finish, add_data=0, new_sheet=1)
    for values in data:
        sheet.append(values)
        row_new += 1
        if row_new == 1048576:
            set_border(sheet, f'A{row}:G{row_new}', row, row_new)
            filters.ref = f"A2:G{row_new}"

            book, sheet, filters = add_new_sheet(
                book=book, sheet_name=sheet_name, avd_time=avd_time, time_start=time_start,
                time_finish=time_finish, add_data=0, new_sheet=1)
            row = sheet.max_row
            row_new = row

    # row_new = sheet.max_row
    set_border(sheet, f'A{row}:F{row_new}', row, row_new)

    # filters = sheet.auto_filter
    filters.ref = f"A2:F{row_new}"
    info(f'Добавлены данные с "{time_start}" по "{time_finish}" в файл.', start_dir)
    # return book
    book.save(file_name)
